"""公众号数据导入模块"""

import os
import re
import sqlite3
from datetime import datetime
from bs4 import BeautifulSoup

from database.schema import init_db, DB_PATH
from scraper.utils import normalize_date, classify_source_by_url

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PUBLIC_ACCOUNT_DIR = os.path.join(BASE_DIR, "公众号数据")


def normalize_game_name(name):
    if not isinstance(name, str):
        return ""
    name_orig = name
    name = name.strip()
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'[（(](?:官服|测试|预约|删档|内测|公测|[\d.]+版本?|S\d)[^）)]*[）)]', '', name)
    return name if name else name_orig


def import_public_account_sheet(filepath):
    """导入公众号 xls/xlsx 到 events 表"""
    import html as html_mod
    import pandas as _pd

    def read_html_table(fpath):
        try:
            dfs = _pd.read_html(fpath, encoding="gbk")
            if dfs is not None and len(dfs) > 0:
                tbl = dfs[0]
                bs = BeautifulSoup("<table></table>", "html.parser")
                t = bs.find("table")
                tr = bs.new_tag("tr")
                for c in tbl.columns:
                    td = bs.new_tag("td")
                    td.string = str(c)
                    tr.append(td)
                t.append(tr)
                for _, r in tbl.iterrows():
                    tr = bs.new_tag("tr")
                    for v in r:
                        td = bs.new_tag("td")
                        vstr = str(v) if str(v) != "nan" else ""
                        td.string = vstr
                        tr.append(td)
                    t.append(tr)
                return t
        except Exception:
            pass
        with open(fpath, "rb") as f:
            raw = f.read()
        meta_enc = None
        meta_match = re.search(rb'charset=[\'\"]?([\w-]+)', raw[:2000])
        if meta_match:
            meta_enc = meta_match.group(1).decode("ascii", errors="ignore").lower()
        encodings = ["gb18030", "gbk", "gb2312", "utf-8-sig", "utf-8", "latin-1"]
        if meta_enc and meta_enc not in encodings:
            encodings = [meta_enc] + encodings
        for enc in encodings:
            try:
                text = raw.decode(enc, errors="replace")
                if "\ufffd" in text:
                    continue
                soup = BeautifulSoup(text, "html.parser")
                tbl = soup.find("table")
                if tbl:
                    return tbl
            except Exception:
                continue
        return None

    table = read_html_table(filepath)
    if not table:
        with open(filepath, "rb") as f:
            raw = f.read()
        for enc in ["utf-8-sig", "utf-8", "gb18030", "gbk", "gb2312", "latin-1"]:
            try:
                text = raw.decode(enc, errors="replace")
            except Exception:
                continue
            soup = BeautifulSoup(text, "html.parser")
            framesets = soup.find_all("frameset")
            target_frames = framesets[-1].find_all("frame") if len(framesets) > 1 else soup.find_all("frame")
            for frame in target_frames:
                src = frame.get("src")
                if not src or "tabstrip" in src.lower():
                    continue
                if src.endswith(".htm") or src.endswith(".html"):
                    sheet_path = os.path.join(os.path.dirname(filepath), src)
                    if os.path.exists(sheet_path):
                        table = read_html_table(sheet_path)
                        if table:
                            break
                    sheet_path2 = os.path.join(BASE_DIR, "公众号数据", src)
                    if os.path.exists(sheet_path2):
                        table = read_html_table(sheet_path2)
                        if table:
                            break
            if table:
                break
    if not table and filepath.lower().endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows_data.append([str(c or "") for c in row])
            if headers and rows_data:
                fake_soup = BeautifulSoup("<table></table>", "html.parser")
                table = fake_soup.find("table")
                tr = fake_soup.new_tag("tr")
                for h in headers:
                    td = fake_soup.new_tag("td")
                    td.string = h
                    tr.append(td)
                table.append(tr)
                for row_data in rows_data:
                    tr = fake_soup.new_tag("tr")
                    for c in row_data:
                        td = fake_soup.new_tag("td")
                        td.string = c
                        tr.append(td)
                    table.append(tr)
        except ImportError:
            pass
    if not table:
        print(f"  跳过 {filepath}: 未找到表格")
        return 0

    rows = table.find_all("tr")
    if len(rows) < 2:
        return 0

    header_cells = rows[0].find_all(["td", "th"])
    headers = [c.get_text(strip=True) for c in header_cells]

    col_map = {}
    for i, h in enumerate(headers):
        h_clean = h.replace("\n", "").replace("\r", "").strip()
        if h_clean in ("游戏名称", "游戏名"):
            col_map["game_name"] = i
        elif h_clean == "游戏ID":
            col_map["game_id"] = i
        elif h_clean in ("测试名称", "测试节点", "节点名称"):
            col_map["test_name"] = i
        elif h_clean in ("服务器地区", "测试服务器", "测试服务器地区"):
            col_map["server_region"] = i
        elif h_clean in ("测试类型", "节点类型"):
            col_map["test_type"] = i
        elif h_clean in ("测试时间", "测试日期"):
            col_map["test_time"] = i
        elif h_clean in ("是否正式运营", "正式运营"):
            col_map["is_formal"] = i
        elif h_clean in ("是否需要激活码", "需要激活码", "激活码", "是否需要邀请码", "需要邀请码", "邀请码"):
            col_map["need_code"] = i
        elif h_clean in ("是否删档", "删档"):
            col_map["is_wiped"] = i
        elif h_clean in ("是否推荐", "推荐"):
            col_map["recommend"] = i
        elif h_clean in ("一句话说明", "说明"):
            col_map["log_text"] = i
        elif h_clean in ("测试公告地址", "公告地址", "链接"):
            col_map["link"] = i
        elif h_clean in ("状态",):
            col_map["status"] = i

    if "game_name" not in col_map or "test_time" not in col_map:
        print(f"  跳过 {filepath}: 缺少必要列")
        return 0

    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_PATH)
    count = 0

    for row in rows[1:]:
        cells = row.find_all("td")
        if len(cells) < max(col_map.values()) + 1:
            continue

        def g(key):
            idx = col_map.get(key)
            if idx is None:
                return ""
            return cells[idx].get_text(strip=True).replace("\n", "")

        game_name = g("game_name")
        if not game_name:
            continue

        clean_name = normalize_game_name(game_name)
        test_name = g("test_name") or "其他"
        test_time = normalize_date(g("test_time"))
        test_type = g("test_type") or "其他"

        if not test_time:
            continue

        conn.execute("""
            INSERT INTO games (clean_name, display_name, first_seen)
            VALUES (?, ?, ?)
            ON CONFLICT(clean_name) DO UPDATE SET display_name=excluded.display_name
        """, (clean_name, game_name, today))
        cur = conn.execute("SELECT id FROM games WHERE clean_name=?", (clean_name,))
        game_id = cur.fetchone()[0]

        sheet_status = g("status") or "待审"
        if sheet_status not in ("正常", "新增", "待审", "弃用"):
            sheet_status = "待审"

        link_url = g("link")
        link_source = classify_source_by_url(link_url) or "公众号"

        conn.execute("""
            INSERT INTO events (game_id, test_name, test_time, test_type,
                server_region, is_formal, need_code, is_wiped,
                log_text, link, source, scrape_date, status, source_type)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(game_id, test_name, test_time) DO UPDATE SET
                status=excluded.status, source_type=excluded.source_type, scrape_date=excluded.scrape_date
        """, (
            game_id, test_name, test_time, test_type,
            g("server_region"), g("is_formal"), g("need_code"), g("is_wiped"),
            g("log_text"), link_url, link_source, today, sheet_status, "public_account"
        ))
        count += 1

    conn.commit()
    conn.execute("INSERT INTO public_account_sheets (filename, import_date, record_count) VALUES (?, ?, ?)",
                 (os.path.basename(filepath), today, count))
    conn.commit()
    conn.close()
    return count


def import_all_public_account_sheets():
    if not os.path.isdir(PUBLIC_ACCOUNT_DIR):
        print(f"公众号数据目录不存在: {PUBLIC_ACCOUNT_DIR}")
        return 0

    conn = sqlite3.connect(DB_PATH)
    imported = set()
    for row in conn.execute("SELECT filename FROM public_account_sheets"):
        imported.add(row[0])
    conn.close()

    total = 0
    for root, dirs, files in os.walk(PUBLIC_ACCOUNT_DIR):
        for fname in sorted(files):
            if not (fname.endswith(".xls") or fname.endswith(".xlsx")):
                continue
            if fname.startswith("~$"):
                continue
            if fname in imported:
                print(f"  跳过已导入: {fname}")
                continue
            fpath = os.path.join(root, fname)
            print(f"  导入: {fname} ...")
            n = import_public_account_sheet(fpath)
            if n > 0:
                print(f"    → 导入 {n} 条记录")
                total += n
    return total
